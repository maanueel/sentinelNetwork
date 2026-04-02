from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class ExcelExporter:
    def __init__(self):
        self.output_dir = '/tmp'
        
    def create_report(self, devices, network_stats):
        """Crear reporte Excel"""
        wb = Workbook()
        
        # Hoja 1: Dispositivos
        ws_devices = wb.active
        ws_devices.title = "Dispositivos"
        self._create_devices_sheet(ws_devices, devices)
        
        # Hoja 2: Estadísticas de Red
        ws_stats = wb.create_sheet("Estadísticas de Red")
        self._create_stats_sheet(ws_stats, network_stats)
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_red_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def _create_devices_sheet(self, ws, devices):
        """Crear hoja de dispositivos"""
        # Encabezados
        headers = ['IP', 'MAC', 'Hostname', 'Fabricante', 'Modelo', 
                   'CPU (%)', 'RAM (%)', 'Disco (%)', 'Red (MB)', 'Estado']
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row = 2
        for ip, device in devices.items():
            ws.cell(row=row, column=1).value = device.get('ip', '')
            ws.cell(row=row, column=2).value = device.get('mac', '')
            ws.cell(row=row, column=3).value = device.get('hostname', '')
            ws.cell(row=row, column=4).value = device.get('vendor', '')
            ws.cell(row=row, column=5).value = device.get('model', '')
            ws.cell(row=row, column=6).value = round(device.get('cpu_usage', 0), 2)
            ws.cell(row=row, column=7).value = round(device.get('ram_usage', 0), 2)
            ws.cell(row=row, column=8).value = round(device.get('disk_usage', 0), 2)
            ws.cell(row=row, column=9).value = round(device.get('network_usage', 0), 2)
            ws.cell(row=row, column=10).value = 'Activo' if device.get('is_reachable') else 'Inactivo'
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_stats_sheet(self, ws, stats):
        """Crear hoja de estadísticas"""
        ws.cell(row=1, column=1).value = "Reporte de Red"
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        ws.cell(row=3, column=1).value = "Fecha:"
        ws.cell(row=3, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.cell(row=5, column=1).value = "Uso de Ancho de Banda (Mbps):"
        ws.cell(row=5, column=2).value = stats.get('bandwidth_usage', 0)
        
        ws.cell(row=6, column=1).value = "Red Saturada:"
        ws.cell(row=6, column=2).value = "SÍ" if stats.get('is_saturated') else "NO"
        
        ws.cell(row=7, column=1).value = "Total de Dispositivos:"
        ws.cell(row=7, column=2).value = stats.get('total_devices', 0)
        
        ws.cell(row=8, column=1).value = "Dispositivos Activos:"
        ws.cell(row=8, column=2).value = stats.get('active_devices', 0)
        
        if stats.get('top_consumer'):
            ws.cell(row=10, column=1).value = "Dispositivo con Mayor Consumo:"
            ws.cell(row=10, column=1).font = Font(bold=True)
            
            top = stats['top_consumer']
            ws.cell(row=11, column=1).value = "IP:"
            ws.cell(row=11, column=2).value = top.get('ip', '')
            ws.cell(row=12, column=1).value = "Hostname:"
            ws.cell(row=12, column=2).value = top.get('hostname', '')
            ws.cell(row=13, column=1).value = "CPU (%):"
            ws.cell(row=13, column=2).value = top.get('cpu', 0)
            ws.cell(row=14, column=1).value = "RAM (%):"
            ws.cell(row=14, column=2).value = top.get('ram', 0)